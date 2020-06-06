"""
One factor hull white model
4/7/2019
"""
import openpyxl
import pandas as pd
import datetime
from scipy.interpolate import interp1d
import numpy as np
# wb = openpyxl.load_workbook(filename = r'D:\Carrie\20Y Swap Rate simulation update weekend.xlsm', data_only=True)
# ws = wb.active
#
# def getTimeFromStr(string):
#     try:
#         return datetime.datetime.strptime(string, '%m%d%Y')
#     except:
#         return datetime.datetime.strptime(string, '%Y-%m-%d %H:%M:%S')
#
# meanReversionRate = ws['B6'].value
# shortRateVol = ws['B7'].value
#
# timeInput = [datetime.datetime.strptime(str(i[0].value), '%m%d%Y') for i in list(ws['$AB$12':'$AB$45'])]
# dfInput = [i[0].value for i in ws['$AC$12':'$AC$45']]
# dtInput = [0] * len(timeInput)
# dtInput =
# startDate = datetime.datetime(2008, 3, 6).date()

meanReversionRate = 0.03
shortRateVol = 0.00726
dateInput = [
'03/06/2008',
'06/10/2008',
'06/18/2008',
'09/17/2008',
'12/17/2008',
'03/18/2009',
'06/17/2009',
'09/16/2009',
'03/10/2010',
'03/10/2011',
'03/12/2012',
'03/11/2013',
'03/10/2014',
'03/10/2015',
'03/10/2016',
'03/10/2017',
'03/12/2018',
'03/11/2019',
'03/10/2020',
'03/10/2023',
'03/10/2028',
'03/10/2033',
'03/10/2038',
'03/10/2048',
'03/11/2058'
]
discountInput = [
1,
0.992416833,
0.992058718,
0.985943647,
0.980166813,
0.974357258,
0.968371794,
0.962046751,
0.949062936,
0.915357101,
0.876387981,
0.833965217,
0.792097728,
0.74987035,
0.709127112,
0.669579432,
0.631786236,
0.595940462,
0.561006901,
0.468428146,
0.352154266,
0.269066757,
0.206770373,
0.12671633,
0.091786689
]
# convert from string to datetime.date object
dateInput = [datetime.datetime.strptime(i, '%m/%d/%Y').date() for i in dateInput]
dtInput = [0] * len(dateInput)
for i in range(1, len(dateInput)):
    # calculate date difference
    dtInput[i] = (dateInput[i] - dateInput[0]).days/365.0


# creating pricing date
startDate = datetime.datetime(2008, 3, 25).date()
simulationDate = [datetime.datetime(2008, 3, 6).date(), startDate]

date = startDate
while date <= datetime.datetime(2018, 4, 10).date():
    date = date + datetime.timedelta(days=1)
    if date.isoweekday() != 6 and date.isoweekday() != 7:
        simulationDate.append(date)

# t and dt
tList = [(date - simulationDate[0]).days/365.0 for date in simulationDate]
dtList = [0] * len(tList)
for i in range(0,len(dtList)-1):
    dtList[i] = tList[i+1] - tList[i]

# interpolation
f = interp1d(dtInput, dtInput)
interpolatedDF = [f(t) for t in tList]
forwardList = [0] *len(interpolatedDF)
for i in range(0, len(forwardList)-1):
    forwardList[i] = (np.log(interpolatedDF[i]) - np.log(interpolatedDF[i+1]))/dtList[i]